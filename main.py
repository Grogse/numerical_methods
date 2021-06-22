import numpy as np
import numpy.linalg as lin
from linear_systems.precise_methods import lu, oem, srm, qr, bordering
from linear_systems.iterative_methods import sim, seidel, sor, gradient, richardson

import openpyxl as excel
from openpyxl.styles import Alignment


def pp(l):
    return '[ ' + ', '.join(f'{n:.3f}' for n in l) + ' ]'


data = np.loadtxt('input.txt')

wb = excel.Workbook()
ws = wb.active
j = 2

matrix_a = data[:, :len(data[0]) - 1]
vector_b = np.squeeze(np.asarray(data[:, len(data[0]) - 1:]))

x_0 = np.round(lin.inv(matrix_a).dot(vector_b), 7)
x_1 = lu.solve_le(data)
x_2 = oem.solve_le(data)
x_3 = srm.solve_le(data)
x_4 = qr.solve_le(data)
x_5 = bordering.solve_le(data)
x_6 = sim.solve_le(data)
x_7 = seidel.solve_le(data)
x_8 = sor.solve_le(data)
x_9 = gradient.solve_le(data)
x_10 = richardson.solve_le(data)

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=11)
ws.merge_cells(start_row=2, start_column=13, end_row=2, end_column=14)

ws.cell(row=1, column=1).value = 'PRECISE METHODS'
ws.cell(row=j, column=1).value = 'LU'
ws.cell(row=j, column=4).value = 'OEM'
ws.cell(row=j, column=7).value = 'SRM'
ws.cell(row=j, column=10).value = 'QR'
ws.cell(row=j, column=13).value = 'BORDERING'

ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
ws.cell(row=j, column=1).alignment = Alignment(horizontal='center')
ws.cell(row=j, column=4).alignment = Alignment(horizontal='center')
ws.cell(row=j, column=7).alignment = Alignment(horizontal='center')
ws.cell(row=j, column=10).alignment = Alignment(horizontal='center')
ws.cell(row=j, column=13).alignment = Alignment(horizontal='center')

for i, k, z, w, q in zip(x_1, x_2, x_3, x_4, x_5):
    ws.cell(row=j + 1, column=1).value = 'x_' + str(j - 1)
    ws.cell(row=j + 1, column=1).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=4).value = 'x_' + str(j - 1)
    ws.cell(row=j + 1, column=4).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=7).value = 'x_' + str(j - 1)
    ws.cell(row=j + 1, column=7).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=10).value = 'x_' + str(j - 1)
    ws.cell(row=j + 1, column=10).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=13).value = 'x_' + str(j - 1)
    ws.cell(row=j + 1, column=13).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=2).value = i
    ws.cell(row=j + 1, column=2).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=5).value = k
    ws.cell(row=j + 1, column=5).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=8).value = z
    ws.cell(row=j + 1, column=8).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=11).value = w
    ws.cell(row=j + 1, column=11).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 1, column=14).value = q
    ws.cell(row=j + 1, column=14).alignment = Alignment(horizontal='center')
    j += 1

j += 2
k = 1

ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=14)
ws.merge_cells(start_row=j + 1, start_column=1, end_row=j + 1, end_column=2)
ws.merge_cells(start_row=j + 1, start_column=4, end_row=j + 1, end_column=5)
ws.merge_cells(start_row=j + 1, start_column=7, end_row=j + 1, end_column=8)
ws.merge_cells(start_row=j + 1, start_column=10, end_row=j + 1, end_column=11)
ws.merge_cells(start_row=j + 1, start_column=13, end_row=j + 1, end_column=14)

ws.cell(row=j, column=1).value = 'ITERATIVE METHODS'
ws.cell(row=j + 1, column=1).value = 'SIM'
ws.cell(row=j + 1, column=4).value = 'SEIDEL'
ws.cell(row=j + 1, column=7).value = 'SOR'
ws.cell(row=j + 1, column=10).value = 'GRADIENT'
ws.cell(row=j + 1, column=13).value = 'RICHARDSON'

ws.cell(row=j, column=1).alignment = Alignment(horizontal='center')
ws.cell(row=j + 1, column=1).alignment = Alignment(horizontal='center')
ws.cell(row=j + 1, column=4).alignment = Alignment(horizontal='center')
ws.cell(row=j + 1, column=7).alignment = Alignment(horizontal='center')
ws.cell(row=j + 1, column=10).alignment = Alignment(horizontal='center')
ws.cell(row=j + 1, column=13).alignment = Alignment(horizontal='center')

for i, z, w, v, m in zip(x_6, x_7, x_8, x_9, x_10):
    ws.cell(row=j + 2, column=1).value = 'x_' + str(k)
    ws.cell(row=j + 2, column=1).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=4).value = 'x_' + str(k)
    ws.cell(row=j + 2, column=4).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=7).value = 'x_' + str(k)
    ws.cell(row=j + 2, column=7).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=10).value = 'x_' + str(k)
    ws.cell(row=j + 2, column=10).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=13).value = 'x_' + str(k)
    ws.cell(row=j + 2, column=13).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=2).value = i
    ws.cell(row=j + 2, column=2).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=5).value = z
    ws.cell(row=j + 2, column=5).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=8).value = w
    ws.cell(row=j + 2, column=8).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=11).value = v
    ws.cell(row=j + 2, column=11).alignment = Alignment(horizontal='center')

    ws.cell(row=j + 2, column=14).value = m
    ws.cell(row=j + 2, column=14).alignment = Alignment(horizontal='center')

    j += 1
    k += 1

wb.save("answers.xlsx")
wb.close()

print('PRECISE METHODS:', x_0)
print('LU -', x_1)
print('OEM -', x_2)
print('SRM -', x_3)
print('QR -', x_4)
print('BORDERING -', x_5)
print(np.array_equal(x_1, x_2) and np.array_equal(x_2, x_4) and np.array_equal(x_4, x_5)
      and np.array_equal(x_5, x_0))

print('\nITERATIVE METHODS:')
print('SIM -', pp(x_6))
print('SEIDEL -', pp(x_7))
print('SOR -', pp(x_8))
print('GRADIENT -', pp(x_9))
print('RICHARDSON -', pp(x_10))
